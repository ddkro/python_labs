import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использует openpyxl для создания Excel файла.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)

    if not csv_path.exists():
        raise FileNotFoundError(f"Файл {csv_path} не найден")

    with csv_path.open(encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) == 0:
        raise ValueError("CSV файл пустой")

    if xlsx_path.suffix.lower() != '.xlsx':
        raise ValueError("Путь должен иметь расширение .xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    for row_idx, row_data in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, row_data in enumerate(zip(*rows), start=1):
        max_length = max(len(str(cell)) for cell in row_data) if row_data else 0
        adjusted_width = max(max_length + 2, 8)
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(xlsx_path)

csv_to_xlsx("data\samples\cities.csv", "data\out\cities.xlsx")